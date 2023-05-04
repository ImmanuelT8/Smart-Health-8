
import pandas as pd

import asyncio
import threading
from tkinter import *
from bleak import BleakClient
import numpy
from prettytable import PrettyTable
import openpyxl
from datetime import datetime


address = "CC:50:E3:9C:15:02"
MODEL_NBR_UUID = "beb5483e-36e1-4688-b7f5-ea07361b26a8"

table = PrettyTable()
table.field_names = ["Wert 1", "Wert 2"]

wb = openpyxl.Workbook() # Workbook-Objekt erzeugen
ws = wb.active #Arbeitsblatt auswählen

now = datetime.now()
dt_string = now.strftime("%d-%m-%Y_%H-%M-%S") # Variable mit Datum und Uhrzeit erzeugen

filename = f"{dt_string}.xlsx" #Datei mit Name und Datum erzeugen

ws["A1"] = "Red"
ws["B1"] = "IR" #Überschriften hinzufügen

last_red_row = 1  # Die erste Zeile ist bereits mit Überschriften belegt
last_ir_row = 1

spo2_list = []
ir_list = []






def print_table():
    print(table)

class BLEThread(threading.Thread):
    def __init__(self, address):
        threading.Thread.__init__(self)
        self.address = address


    async def handle_uuid1_notify(self, sender, data):
        spo2_values = numpy.frombuffer(data, dtype=numpy.uint32)
        print("Red: {0}".format(spo2_values))
        for spo2_value in spo2_values:
            table.add_row([spo2_value, ""])
            ws.append([spo2_value, ""])  # Zeile hinzufügen
            spo2_list.append(spo2_value)  # füge den empfangenen Wert der Liste hinzu

        # wenn alle Werte empfangen wurden, erstelle einen Pandas DataFrame und gib ihn aus
        if len(spo2_list) == 100:
            pythonDaten = pd.DataFrame(spo2_list)
            print(pythonDaten.head())
            wb.save(filename)


    async def handle_uuid2_notify(self, sender, data):

        ir_values = numpy.frombuffer(data, dtype=numpy.uint32)
        print("IR: {0}".format(ir_values))
        for ir_value in ir_values:
            if ir_value:
                table.add_row(["", ir_value])
                ws.append(["", ir_value])  # Zeile hinzufügen
                ir_list.append(ir_value)  # füge den empfangenen Wert der Liste hinzu


        # wenn alle Werte empfangen wurden, erstelle einen Pandas DataFrame und gib ihn aus
        if len(ir_list) == 100:
            pythonDaten2 = pd.DataFrame(spo2_list)
            print(pythonDaten2.head())
            wb.save(filename)

    async def read_data(self):
        UUID1 = "beb5483e-36e1-4688-b7f5-ea07361b26a8"
        UUID2 = "beb5483e-36e1-4688-b7f5-ea07361b26a9"
        async with BleakClient(self.address) as client:
            await client.start_notify(UUID1, self.handle_uuid1_notify)
            await client.start_notify(UUID2, self.handle_uuid2_notify)
            while True:
                await asyncio.sleep(1)

    def run(self):
        asyncio.run(self.read_data())


class TkinterThread(threading.Thread):
    def __init__(self):
        threading.Thread.__init__(self)

    def run(self):
        tkFenster = Tk()
        tkFenster.title('BLE MAX30102')
        tkFenster.mainloop()


def start_threads():
    ble_thread = BLEThread(address)
    tkinter_thread = TkinterThread()

    ble_thread.start()
    tkinter_thread.start()


if __name__ == '__main__':
    start_threads()

wb.save(filename)
