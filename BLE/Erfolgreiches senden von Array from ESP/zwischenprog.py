import asyncio
import threading
from tkinter import *
import openpyxl
import datetime

from bleak import BleakClient
import numpy
from prettytable import PrettyTable

# Bluetooth Adress of the device
address = "CC:50:E3:9C:15:02"
# UUID of the service to read data from
UUID = "beb5483e-36e1-4688-b7f5-ea07361b26a8"

# Create a new workbook and worksheet in the same directory
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "BLE Data"

# Set the headers for the columns in the worksheet
worksheet.cell(row=1, column=1, value="Wert 1")
worksheet.cell(row=1, column=2, value="Wert 2")

# Initialize the PrettyTable to display the data in the console
table = PrettyTable()
table.field_names = ["Wert 1", "Wert 2"]


def print_table():
    """
    Function to print the PrettyTable to the console
    """
    print(table)


class BLEThread(threading.Thread):
    """
    Thread to read data from the Bluetooth device
    """
    def __init__(self, address):
        threading.Thread.__init__(self)
        self.address = address

    async def handle_notify(self, sender, data):
        """
        Callback function to handle data received from the device
        """
        values = numpy.frombuffer(data, dtype=numpy.uint32)
        print("Received values: {0}".format(values))

        for i in range(0, len(values), 2):
            # Add the received values to the worksheet
            worksheet.append([values[i], values[i+1]])
            # Add the received values to the PrettyTable
            table.add_row([values[i], values[i+1]])

        # Print the table to the console
        print_table()

    async def read_data(self):
        """
        Function to connect to the device and start reading data
        """
        async with BleakClient(self.address) as client:
            await client.start_notify(UUID, self.handle_notify)
            while True:
                await asyncio.sleep(1)

    def run(self):
        asyncio.run(self.read_data())


class TkinterThread(threading.Thread):
    """
    Thread to run the Tkinter mainloop
    """
    def __init__(self):
        threading.Thread.__init__(self)

    def run(self):
        tkFenster = Tk()
        tkFenster.title('BLE MAX30102')
        tkFenster.mainloop()


def save_table():
    """
    Function to save the data in the worksheet to an Excel file
    """
    now = datetime.datetime.now()
    filename = "BLE_data_{}.xlsx".format(now.strftime("%Y-%m-%d_%H-%M-%S"))
    workbook.save(filename)


def start_threads():
    """
    Function to start the BLE and Tkinter threads
    """
    ble_thread = BLEThread(address)
    tkinter_thread = TkinterThread()

    ble_thread.start()
    tkinter_thread.start()
    threading.Timer(60.0, save_table).start()


if __name__ == '__main__':
    start_threads()import asyncio
import threading
from tkinter import *
import openpyxl
import datetime

from bleak import BleakClient
import numpy
from prettytable import PrettyTable

# Bluetooth Adress of the device
address = "CC:50:E3:9C:15:02"
# UUID of the service to read data from
UUID = "beb5483e-36e1-4688-b7f5-ea07361b26a8"

# Create a new workbook and worksheet in the same directory
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "BLE Data"

# Set the headers for the columns in the worksheet
worksheet.cell(row=1, column=1, value="Wert 1")
worksheet.cell(row=1, column=2, value="Wert 2")

# Initialize the PrettyTable to display the data in the console
table = PrettyTable()
table.field_names = ["Wert 1", "Wert 2"]


def print_table():
    """
    Function to print the PrettyTable to the console
    """
    print(table)


class BLEThread(threading.Thread):
    """
    Thread to read data from the Bluetooth device
    """
    def __init__(self, address):
        threading.Thread.__init__(self)
        self.address = address

    async def handle_notify(self, sender, data):
        """
        Callback function to handle data received from the device
        """
        values = numpy.frombuffer(data, dtype=numpy.uint32)
        print("Received values: {0}".format(values))

        for i in range(0, len(values), 2):
            # Add the received values to the worksheet
            worksheet.append([values[i], values[i+1]])
            # Add the received values to the PrettyTable
            table.add_row([values[i], values[i+1]])

        # Print the table to the console
        print_table()

    async def read_data(self):
        """
        Function to connect to the device and start reading data
        """
        async with BleakClient(self.address) as client:
            await client.start_notify(UUID, self.handle_notify)
            while True:
                await asyncio.sleep(1)

    def run(self):
        asyncio.run(self.read_data())


class TkinterThread(threading.Thread):
    """
    Thread to run the Tkinter mainloop
    """
    def __init__(self):
        threading.Thread.__init__(self)

    def run(self):
        tkFenster = Tk()
        tkFenster.title('BLE MAX30102')
        tkFenster.mainloop()


def save_table():
    """
    Function to save the data in the worksheet to an Excel file
    """
    now = datetime.datetime.now()
    filename = "BLE_data_{}.xlsx".format(now.strftime("%Y-%m-%d_%H-%M-%S"))
    workbook.save(filename)


def start_threads():
    """
    Function to start the BLE and Tkinter threads
    """
    ble_thread = BLEThread(address)
    tkinter_thread = TkinterThread()

    ble_thread.start()
    tkinter_thread.start()
    threading.Timer(60.0, save_table).start()


if __name__ == '__main__':
    start_threads()
